#!/usr/bin/env python3
"""
Main entrypoint
"""
import click
from tabulate import tabulate
import openpyxl
from openpyxl.styles import Font
from gpt.ask import query_gpt
from ec2.ec2scan import query_ec2
from ebs.ebsscan import query_ebs, query_ebs_snapshots
from rds.rdsscan import query_rds
from ami.amiscan import query_ami
from ecr.ecrscan import query_ecr_images
from lb.lbscan import query_lb
from cloudwatch.group import query_cloudwatch_groups


def export_data(workbook, query_func, region, table_head, table_data):
    """
    Export data
    """

    sheet_name = f"{region}_{query_func.__name__}"[:30]
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(table_head)

    bold_font = Font(bold=True)
    for col_num in range(1, len(table_head) + 1):
        sheet.cell(row=1, column=col_num).font = bold_font

    for row in table_data:
        sheet.append(row)


def tabulate_data(ai, mode, table_head, table_data):
    """
    Tabulate data
    """

    tabulated_data = tabulate(
        table_data, headers=table_head, tablefmt="github", floatfmt=".2f"
    )
    if len(table_data) < 2000:
        print(tabulated_data)
    else:
        print("❗ Too many rows to print, use --export-file to export to Excel")

    if ai:
        query_gpt(mode, table_head, tabulated_data)


@click.command(context_settings={"show_default": True})
@click.help_option("-h", "--help")
@click.option(
    "-m",
    "--modes",
    required=False,
    default="ebs,ec2,rds,lb,ami,ecr,cw",
)
@click.option(
    "-r",
    "--regions",
    required=True,
)
@click.option(
    "-a",
    "--ai-suggestions",
    is_flag=True,
    default=False,
)
@click.option(
    "-e",
    "--export-file",
    help="Export file path",
    required=False,
    default="costs-optimizer.xlsx",
)
def main(**options):
    """
    Main entrypoint
    """

    modes = options["modes"]
    regions = options["regions"]
    ai = options["ai_suggestions"]
    export_file = options["export_file"]

    workbook = openpyxl.Workbook()
    workbook.remove(workbook["Sheet"])

    mode_functions = {
        "ebs": [query_ebs, query_ebs_snapshots],
        "ec2": [query_ec2],
        "rds": [query_rds],
        "lb": [query_lb],
        "ami": [query_ami],
        "ecr": [query_ecr_images],
        "cw": [query_cloudwatch_groups],
    }

    for region in regions.split(","):
        for mode in modes.split(","):
            if mode in mode_functions:
                for query_func in mode_functions[mode]:
                    table_head, table_data = query_func(region)
                    if table_head and table_data:
                        tabulate_data(ai, mode, table_head, table_data)

                    if export_file:
                        export_data(
                            workbook, query_func, region, table_head, table_data
                        )

    workbook.save(export_file)


if __name__ == "__main__":
    main()
